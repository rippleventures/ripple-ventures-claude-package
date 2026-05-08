#!/usr/bin/env python3
"""
extract_named_ranges.py — Extract all named ranges from an Excel workbook.

Used by build-deck and other skills that bind PowerPoint placeholders to workbook data.
Each named range becomes a key the deck-build script can look up.

Usage:
  python extract_named_ranges.py path/to/workbook.xlsx

Output: JSON map of {name: {sheet, range, value}} to stdout.

Dependencies: openpyxl
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed", "fix": "pip install openpyxl"}))
    sys.exit(2)


def extract_named_ranges(path):
    wb = load_workbook(path, data_only=True)  # data_only=True returns evaluated values
    result = {}

    for defined_name in wb.defined_names:
        try:
            destinations = list(wb.defined_names[defined_name].destinations)
        except KeyError:
            continue

        for sheet_name, range_str in destinations:
            try:
                ws = wb[sheet_name]
                cell_range = ws[range_str]

                # Single cell
                if not isinstance(cell_range, tuple):
                    value = cell_range.value
                else:
                    # Multi-cell range — flatten to list
                    value = []
                    for row in cell_range:
                        if isinstance(row, tuple):
                            value.append([c.value for c in row])
                        else:
                            value.append(row.value)

                result[defined_name] = {
                    "sheet": sheet_name,
                    "range": range_str,
                    "value": value,
                }
            except (KeyError, AttributeError) as e:
                result[defined_name] = {
                    "sheet": sheet_name,
                    "range": range_str,
                    "error": str(e),
                }

    return result


def main():
    parser = argparse.ArgumentParser(description="Extract named ranges from xlsx")
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()

    if not args.workbook.exists():
        print(json.dumps({"error": f"File not found: {args.workbook}"}))
        sys.exit(2)

    ranges = extract_named_ranges(args.workbook)
    print(json.dumps(ranges, indent=2, default=str))


if __name__ == "__main__":
    main()
