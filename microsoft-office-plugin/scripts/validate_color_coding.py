#!/usr/bin/env python3
"""
validate_color_coding.py — Verify Excel color coding follows the institutional standard.

Standard:
  - Blue (0070C0) — typed inputs (numbers, text)
  - Black (000000) — formulas referencing only the same sheet
  - Green (00B050) — formulas referencing other sheets
  - Red (FF0000) — formulas linking to external workbooks (should be removed)

Usage:
  python validate_color_coding.py path/to/workbook.xlsx

Output: JSON report. Lists any cells that violate the convention.

Dependencies: openpyxl
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed", "fix": "pip install openpyxl"}))
    sys.exit(2)


COLOR_BLUE = "0070C0"
COLOR_BLACK = "000000"
COLOR_GREEN = "00B050"
COLOR_RED = "FF0000"


def font_color(cell):
    if cell.font and cell.font.color and cell.font.color.value:
        val = cell.font.color.value
        if isinstance(val, str):
            return val.upper().lstrip("FF")
    return None


def is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")


def is_cross_sheet_formula(cell, current_sheet):
    if not is_formula(cell):
        return False
    # Look for SheetName! reference NOT preceded by a [ (which would indicate external link)
    pattern = re.compile(r"(?<!\[)([A-Za-z_][\w\s]*)!", re.UNICODE)
    matches = pattern.findall(cell.value)
    return any(m.strip("' ").lower() != current_sheet.lower() for m in matches)


def is_external_link(cell):
    return is_formula(cell) and "[" in cell.value and ".xls" in cell.value.lower()


def is_typed_value(cell):
    if cell.value is None:
        return False
    return not is_formula(cell) and cell.value != ""


def expected_color(cell, sheet_name):
    if is_external_link(cell):
        return COLOR_RED
    if is_cross_sheet_formula(cell, sheet_name):
        return COLOR_GREEN
    if is_formula(cell):
        return COLOR_BLACK
    if is_typed_value(cell):
        return COLOR_BLUE
    return None  # empty


def validate(path):
    wb = load_workbook(path, data_only=False)
    violations = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                expected = expected_color(cell, sheet_name)
                if expected is None:
                    continue
                actual = font_color(cell)
                if actual != expected:
                    violations.append({
                        "cell": f"{sheet_name}!{cell.coordinate}",
                        "expected_color": expected,
                        "actual_color": actual or "default",
                        "value_type": (
                            "external_link" if is_external_link(cell)
                            else "cross_sheet_formula" if is_cross_sheet_formula(cell, sheet_name)
                            else "formula" if is_formula(cell)
                            else "typed_value"
                        ),
                    })

    return {
        "workbook": str(path),
        "passed": len(violations) == 0,
        "violation_count": len(violations),
        "violations": violations,
    }


def main():
    parser = argparse.ArgumentParser(description="Validate Excel color coding")
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()

    if not args.workbook.exists():
        print(json.dumps({"error": f"File not found: {args.workbook}"}))
        sys.exit(2)

    report = validate(args.workbook)
    print(json.dumps(report, indent=2))

    if not report["passed"]:
        sys.exit(1)


if __name__ == "__main__":
    main()
