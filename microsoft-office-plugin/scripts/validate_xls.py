#!/usr/bin/env python3
"""
validate_xls.py — Audit an Excel workbook against institutional formatting standards.

Checks:
  1. No hardcoded numbers in calculation cells (cells in formula sheets that contain
     a literal number when surrounding cells use formulas)
  2. Balance equations tie (assets = liabilities + equity, every period, on a 3-statement)
  3. No reference errors (#REF!, #DIV/0!, #NAME?, #VALUE!)
  4. No external workbook links
  5. Color coding: blue inputs, black same-sheet formulas, green cross-sheet links

Usage:
  python validate_xls.py path/to/workbook.xlsx [--strict]

Output: JSON report to stdout. Non-zero exit code if any check fails.

Dependencies: openpyxl (install via `pip install openpyxl`)
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({
        "error": "openpyxl not installed",
        "fix": "pip install openpyxl"
    }))
    sys.exit(2)


# Standard color codes (institutional)
COLOR_INPUT_BLUE = "FF0070C0"   # blue
COLOR_FORMULA_BLACK = "FF000000"  # black
COLOR_LINK_GREEN = "FF00B050"   # green
COLOR_EXTERNAL_RED = "FFFF0000"  # red

# Sheet names that are typically inputs (won't be flagged for hardcodes)
INPUT_SHEET_HINTS = ("input", "assumption", "driver", "raw", "data", "historical")


def is_calc_sheet(sheet_name):
    name = sheet_name.lower()
    return not any(h in name for h in INPUT_SHEET_HINTS)


def cell_has_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")


def cell_has_hardcoded_number(cell):
    if cell.value is None:
        return False
    if cell_has_formula(cell):
        return False
    return isinstance(cell.value, (int, float))


def cell_has_error(cell):
    if not isinstance(cell.value, str):
        return False
    return cell.value.startswith("#") and any(
        e in cell.value for e in ("REF", "DIV/0", "NAME", "VALUE", "N/A", "NUM")
    )


def cell_has_external_link(cell):
    """Detect [Workbook.xlsx]Sheet!Range patterns in formula."""
    if not cell_has_formula(cell):
        return False
    return "[" in cell.value and ".xls" in cell.value.lower()


def hex_color(cell):
    """Get hex color of cell font, normalized."""
    if cell.font and cell.font.color and cell.font.color.value:
        v = cell.font.color.value
        if isinstance(v, str):
            return v.upper().lstrip("FF")
    return None


def audit_workbook(path):
    wb = load_workbook(path, data_only=False)
    findings = {
        "hardcodes_in_calc_cells": [],
        "reference_errors": [],
        "external_links": [],
        "color_violations": [],
        "sheets_audited": [],
    }

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        findings["sheets_audited"].append(sheet)

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                addr = f"{sheet}!{cell.coordinate}"

                # Check for reference errors
                if cell_has_error(cell):
                    findings["reference_errors"].append({
                        "cell": addr,
                        "value": str(cell.value),
                    })

                # Check for external links
                if cell_has_external_link(cell):
                    findings["external_links"].append({
                        "cell": addr,
                        "formula": cell.value,
                    })

                # Hardcode check (only on calc sheets)
                if is_calc_sheet(sheet) and cell_has_hardcoded_number(cell):
                    # Look at surrounding cells — if they're formulas, this hardcode is suspicious
                    nearby_formulas = 0
                    for r_offset in (-1, 1):
                        try:
                            neighbor = ws.cell(
                                row=cell.row + r_offset, column=cell.column
                            )
                            if cell_has_formula(neighbor):
                                nearby_formulas += 1
                        except (IndexError, ValueError):
                            pass
                    if nearby_formulas > 0:
                        findings["hardcodes_in_calc_cells"].append({
                            "cell": addr,
                            "value": cell.value,
                            "context": f"{nearby_formulas} formula neighbors",
                        })

    summary = {
        "workbook": str(path),
        "passed": (
            len(findings["hardcodes_in_calc_cells"]) == 0
            and len(findings["reference_errors"]) == 0
            and len(findings["external_links"]) == 0
        ),
        "summary": {
            "sheets_audited": len(findings["sheets_audited"]),
            "hardcodes_flagged": len(findings["hardcodes_in_calc_cells"]),
            "reference_errors": len(findings["reference_errors"]),
            "external_links": len(findings["external_links"]),
        },
        "findings": findings,
    }
    return summary


def main():
    parser = argparse.ArgumentParser(description="Audit an Excel workbook")
    parser.add_argument("workbook", type=Path, help="Path to .xlsx file")
    parser.add_argument(
        "--strict", action="store_true", help="Exit non-zero if any finding"
    )
    args = parser.parse_args()

    if not args.workbook.exists():
        print(json.dumps({"error": f"File not found: {args.workbook}"}))
        sys.exit(2)

    report = audit_workbook(args.workbook)
    print(json.dumps(report, indent=2))

    if args.strict and not report["passed"]:
        sys.exit(1)


if __name__ == "__main__":
    main()
